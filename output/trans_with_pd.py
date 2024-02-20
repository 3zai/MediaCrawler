import json
import glob
import pandas as pd
import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from tqdm import tqdm
from io import BytesIO

file_list = glob.glob("../data/xhs/*contents*.json")
sheet_list = []
key_list = ['note_id', 'title', 'desc', 'time', 'user_id', 'nickname', 'liked_count', 'collected_count',
            'comment_count', 'share_count', 'image_list']
target_file = '333.xlsx'
img_path_index = 0
# img_path = '000.jpg'


def download(url, path):
    if os.path.exists(path):
        os.remove(path)
    file = requests.get(url)
    if file.status_code != 200:
        return
    img = PILImage.open(BytesIO(file.content))
    img = img.convert("RGB")  # 下载的数据为png格式,除了rgb三色信息外还包含了透明度信息,无法保存为jpg格式
    img.save(path)
    # with open(path, 'wb') as img:
    #     img.write(file.content)

def get_cell_info(cell_name):
    left = ""
    right = ""

    for i in cell_name:
        if not i.isnumeric():
            left += i
        else:
            right += i
    return int(right), left


def run():
    global img_path_index
    if os.path.exists(target_file):
        os.remove(target_file)

    for i in range(len(file_list)):
        sheet_list.append(f"Sheet{i}")
        with open(file_list[i], 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        target_map = {}
        for j in range(len(json_data)):
            for key, value in json_data[j].items():
                if key in key_list:
                    target_map.setdefault(key, [])
                    if not value:
                        print(f"cur key: {key} 's val is null")
                        value = ""
                    target_map[key].append(value)
        df = pd.DataFrame(target_map)
        # 1.按照四种data数据之和进行倒序排序
        df['tmp'] = df['liked_count'] + df['collected_count'] + df['comment_count'] + df['share_count']
        df = df.sort_values(by='tmp', ascending=False)
        df = df.drop('tmp', axis=1)
        # 2.image_list在excel内拆分
        for index in range(len(df['image_list'])):
            if ',' not in df['image_list'][index]:
                df[f'image_list_{0}'][index] = df['image_list'][index]
            else:
                img_list = df['image_list'][index].split(',')
                for k in range(len(img_list)):
                    if f'image_list_{k}' not in df.columns:
                        df[f'image_list_{k}'] = None
                    df[f'image_list_{k}'][index] = img_list[k]
        # 3.导出excel, todo 当前是覆盖式的, 需要修改为新增表
        if not os.path.exists(target_file):
            df.to_excel(target_file, sheet_name=f"Sheet{i}", index=False, na_rep="")
        else:
            with pd.ExcelWriter(target_file, engine='openpyxl', if_sheet_exists='replace', mode='a') as writer:
                df.to_excel(target_file, sheet_name=f"Sheet{i}", index=False, na_rep="")
        # 4.遍历image_list, 将链接转换为图片
        book = load_workbook(target_file)
        writer = pd.ExcelWriter(target_file, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        ws = writer.sheets[f'Sheet{i}']

        for j in tqdm(range(len(ws[1]))):
        # for j in tqdm(range(13,15)):
            limit = 2
            cell = ws[1][j]
            if cell.value.startswith('image_list_') and limit > 0:
                limit -= 1
                index = j + 1
                for col in ws.iter_cols(min_col=index, max_col=index, min_row=2):
                    for k in tqdm(range(len(col))):
                    # for k in tqdm(range(5)):
                        cur_row = 2 + k
                        cur_col = index
                        cur_val = col[k].value
                        if not cur_val:
                            continue
                        img_path = str(img_path_index) + ".jpg"
                        img_path_index += 1
                        download(cur_val, img_path)

                        if os.path.exists(img_path):
                            # 插入图片
                            try:
                                img = PILImage.open(img_path)
                                pic = Image(img)
                                pic.width = 200
                                pic.height = 200
                                cell_name = ws.cell(row=cur_row, column=cur_col).coordinate
                                pic.anchor = cell_name
                                row_name, col_name = get_cell_info(cell_name)

                                ws.column_dimensions[col_name].width = 25
                                ws.row_dimensions[row_name].height = 200
                                ws.add_image(pic)

                            except Exception as e:
                                print(f"[warn] catch image error , {str(e)}")
                                continue
        writer.save()
        image_list = glob.glob("*.jpg")
        for f in image_list:
            os.remove(f)

if __name__ == "__main__":
    run()