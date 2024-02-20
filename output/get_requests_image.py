import xlwt
import openpyxl
import requests
import sys

if "__name__" == "__main__":
    # path = sys.argv[1]
    path = '333.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    for i in range(len(ws[1])):
        cell = ws[1][i]
        if cell.value.startswith('image_list_'):
            index = i + 1
            for col in ws.iter_cols(min_col=index, max_col=index, min_row=1):
                for j in range(len(col)):
                    cur_row = 1 + j
                    cur_col = index
                    cur_val = col[j].value
